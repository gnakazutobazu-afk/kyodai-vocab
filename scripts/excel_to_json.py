#!/usr/bin/env python3
"""
Excel（京大英語完全対策英単語リスト）→ words.json 変換スクリプト

使用例:
    python scripts/excel_to_json.py \
      --input "/Users/nakaz/Library/Mobile Documents/com~apple~CloudDocs/001_ビジネス/109_note/京大英語過去問/京大英語完全対策英単語リスト_fix01.xlsx" \
      --output ./public/data/words.json
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SHEET_NAME = "完全対策リスト"

# word_note から各ブロックを抽出する正規表現
NOTE_PATTERNS = {
    "etymology": re.compile(r"【語源】\s*(.*?)(?=【|$)", re.DOTALL),
    "synonyms": re.compile(r"【類義語】\s*(.*?)(?=【|$)", re.DOTALL),
    "antonyms": re.compile(r"【対義語】\s*(.*?)(?=【|$)", re.DOTALL),
    "kyodai_usage": re.compile(r"【京大での使われ方】\s*(.*?)(?=【|$)", re.DOTALL),
    "memory_tip": re.compile(r"【覚え方】\s*(.*?)(?=【|$)", re.DOTALL),
}

# 例文先頭の出典プレフィックス [YYYY] または (YYYY) を除去
CITATION_RE = re.compile(r"^[\[\(](\d{4})[\]\)]\s*")


def extract_note_blocks(note_text: str) -> dict:
    """word_note から各ブロックを抽出する。"""
    if not note_text:
        return {k: "" for k in NOTE_PATTERNS}
    result = {}
    for key, pattern in NOTE_PATTERNS.items():
        m = pattern.search(note_text)
        result[key] = m.group(1).strip() if m else ""
    return result


def parse_example(raw: str) -> tuple[str, str]:
    """
    例文から出典年度を抽出し (example_en, example_citation) を返す。
    例: "[2000] You had to be there..." → ("You had to be there...", "2000")
    """
    if not raw:
        return "", ""
    m = CITATION_RE.match(raw)
    if m:
        year = m.group(1)
        text = raw[m.end():].strip()
        return text, year
    return raw.strip(), ""


def cell_value(cell) -> str:
    """セルの値を文字列に変換。None は空文字。"""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def compute_level(frequency: int, cefr: str) -> str:
    """出現年度数とCEFRから頻出レベルを判定する。"""
    if frequency >= 3:
        return "頻出"
    if frequency == 2:
        return "重要"
    if frequency == 1 and cefr in ("C1", "B2"):
        return "重要"
    return "発展"


def load_sheet(path: str) -> openpyxl.worksheet.worksheet.Worksheet:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        print(
            f"シート「{SHEET_NAME}」が見つかりません。利用可能なシート: {available}",
            file=sys.stderr,
        )
        sys.exit(1)
    return wb[SHEET_NAME]


def build_words(ws) -> list[dict]:
    rows = list(ws.iter_rows())
    if not rows:
        print("シートが空です。", file=sys.stderr)
        sys.exit(1)

    # 1行目をヘッダーとして取得
    header = [cell_value(c).lower().strip() for c in rows[0]]

    def col(row, name: str) -> str:
        try:
            idx = header.index(name)
            return cell_value(row[idx])
        except ValueError:
            return ""

    words = []
    for i, row in enumerate(rows[1:], start=2):
        word = col(row, "word")
        if not word:
            continue  # word 列が空の行をスキップ

        # 例文処理
        raw_example = col(row, "example_sentence")
        example_en, example_citation = parse_example(raw_example)

        # word_note からブロック抽出
        note_blocks = extract_note_blocks(col(row, "word_note"))

        # has_detail の判定
        has_detail = bool(
            col(row, "example_sentence_ja") and col(row, "word_note")
        )

        # 出現年度数と頻出レベルの計算
        freq_str = col(row, "出現年度数")
        frequency = int(freq_str) if freq_str.isdigit() else 1
        cefr = col(row, "cefr")
        level = compute_level(frequency, cefr)

        entry = {
            "word": word,
            "theme_tag": col(row, "theme_tag"),
            "pos_ja": col(row, "pos_ja"),
            "meaning_ja": col(row, "meaning_ja"),
            "frequency": frequency,
            "level": level,
            "example_en": example_en,
            "example_citation": example_citation,
            "example_ja": col(row, "example_sentence_ja"),
            **note_blocks,
            "has_detail": has_detail,
        }
        words.append(entry)

    return words


def main():
    parser = argparse.ArgumentParser(description="Excel → words.json 変換")
    parser.add_argument("--input", required=True, help="入力 Excel ファイルパス")
    parser.add_argument("--output", required=True, help="出力 JSON ファイルパス")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()

    if not input_path.exists():
        print(f"入力ファイルが見つかりません: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"読み込み中: {input_path}")
    ws = load_sheet(str(input_path))

    print("変換中...")
    words = build_words(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(words, f, ensure_ascii=False, indent=2)

    print(f"完了: {len(words)} 語 → {output_path}")


if __name__ == "__main__":
    main()
